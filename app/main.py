import csv
import io
import json
import os
from datetime import datetime, timezone
from pathlib import Path

from fastapi import Depends, FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.routing import APIRouter
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session

from app.auth import get_current_client
from app.database import get_db, init_db
from app.engine.logic import FORMULA_META, FORMULAS
from app.models import Client, Workbook
from app.schemas import (
    CalculationRequest,
    CalculationResponse,
    ClientInfo,
    LoginRequest,
    WorkbookCreate,
    WorkbookInfo,
    WorkbookSummary,
    WorkbookUpdate,
)

# ─────────────────────────────────────────────────────────────────────────────
# API Router — toutes les routes métier sous /api
# ─────────────────────────────────────────────────────────────────────────────
api = APIRouter(prefix="/api")


@api.get("/health")
def health():
    return {"status": "ok"}


@api.post("/auth/login", response_model=ClientInfo)
def login(payload: LoginRequest, db: Session = Depends(get_db)):
    client = db.query(Client).filter(Client.api_key == payload.api_key).first()
    if client is None:
        raise HTTPException(status_code=401, detail="Clé API invalide.")
    return client


@api.get("/me", response_model=ClientInfo)
def me(client: Client = Depends(get_current_client)):
    return client


@api.get("/formulas")
def list_formulas(client: Client = Depends(get_current_client)):
    return {"formulas": FORMULA_META}


@api.post("/calculate", response_model=CalculationResponse)
def calculate(
    payload: CalculationRequest,
    client: Client = Depends(get_current_client),
    db: Session = Depends(get_db),
):
    formula_fn = FORMULAS.get(payload.formula)
    if formula_fn is None:
        raise HTTPException(
            status_code=404,
            detail=f"Formule '{payload.formula}' introuvable. "
            f"Formules disponibles : {list(FORMULAS.keys())}",
        )

    try:
        result = formula_fn(payload.variables)
    except KeyError as exc:
        raise HTTPException(
            status_code=422,
            detail=f"Variable manquante : {exc}",
        )
    except (TypeError, ValueError) as exc:
        raise HTTPException(
            status_code=422,
            detail=f"Erreur dans les variables fournies : {exc}",
        )

    # Décrémenter les crédits
    client.credits -= 1
    db.commit()
    db.refresh(client)

    return CalculationResponse(
        formula=payload.formula,
        result=result,
        credits_remaining=client.credits,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Workbooks CRUD
# ─────────────────────────────────────────────────────────────────────────────

def _wb_to_dict(wb: Workbook) -> dict:
    return {
        "id": wb.id,
        "name": wb.name,
        "data": json.loads(wb.data),
        "formulas": json.loads(wb.formulas),
        "created_at": wb.created_at.isoformat(),
        "updated_at": wb.updated_at.isoformat(),
    }


def _wb_summary(wb: Workbook) -> dict:
    return {
        "id": wb.id,
        "name": wb.name,
        "created_at": wb.created_at.isoformat(),
        "updated_at": wb.updated_at.isoformat(),
    }


@api.get("/workbooks")
def list_workbooks(
    client: Client = Depends(get_current_client),
    db: Session = Depends(get_db),
):
    wbs = db.query(Workbook).filter(Workbook.client_id == client.id).order_by(Workbook.updated_at.desc()).all()
    return {"workbooks": [_wb_summary(wb) for wb in wbs]}


@api.post("/workbooks", status_code=201)
def create_workbook(
    payload: WorkbookCreate,
    client: Client = Depends(get_current_client),
    db: Session = Depends(get_db),
):
    wb = Workbook(client_id=client.id, name=payload.name)
    db.add(wb)
    db.commit()
    db.refresh(wb)
    return _wb_to_dict(wb)


@api.get("/workbooks/{wb_id}")
def get_workbook(
    wb_id: int,
    client: Client = Depends(get_current_client),
    db: Session = Depends(get_db),
):
    wb = db.query(Workbook).filter(Workbook.id == wb_id, Workbook.client_id == client.id).first()
    if not wb:
        raise HTTPException(status_code=404, detail="Classeur introuvable.")
    return _wb_to_dict(wb)


@api.put("/workbooks/{wb_id}")
def update_workbook(
    wb_id: int,
    payload: WorkbookUpdate,
    client: Client = Depends(get_current_client),
    db: Session = Depends(get_db),
):
    wb = db.query(Workbook).filter(Workbook.id == wb_id, Workbook.client_id == client.id).first()
    if not wb:
        raise HTTPException(status_code=404, detail="Classeur introuvable.")
    if payload.name is not None:
        wb.name = payload.name
    if payload.data is not None:
        wb.data = json.dumps(payload.data)
    if payload.formulas is not None:
        wb.formulas = json.dumps(payload.formulas)
    wb.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(wb)
    return _wb_to_dict(wb)


@api.delete("/workbooks/{wb_id}", status_code=204)
def delete_workbook(
    wb_id: int,
    client: Client = Depends(get_current_client),
    db: Session = Depends(get_db),
):
    wb = db.query(Workbook).filter(Workbook.id == wb_id, Workbook.client_id == client.id).first()
    if not wb:
        raise HTTPException(status_code=404, detail="Classeur introuvable.")
    db.delete(wb)
    db.commit()


@api.post("/workbooks/import")
async def import_workbook(
    file: UploadFile = File(...),
    client: Client = Depends(get_current_client),
    db: Session = Depends(get_db),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Nom de fichier manquant.")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    raw = await file.read()

    if ext == "csv":
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
    elif ext in ("xlsx", "xls"):
        try:
            import openpyxl
            wb_xl = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            ws = wb_xl.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
        except ImportError:
            raise HTTPException(status_code=400, detail="Import XLSX non supporte (openpyxl manquant).")
    else:
        raise HTTPException(status_code=400, detail=f"Format '{ext}' non supporte. Utilisez CSV ou XLSX.")

    cols = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    cell_data: dict[str, dict[str, str]] = {}
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row):
            if c_idx >= 26:
                break
            col_letter = cols[c_idx]
            if val.strip():
                if col_letter not in cell_data:
                    cell_data[col_letter] = {}
                cell_data[col_letter][str(r_idx)] = val.strip()

    name = file.filename.rsplit(".", 1)[0] if "." in file.filename else file.filename
    wb = Workbook(
        client_id=client.id,
        name=name,
        data=json.dumps(cell_data),
    )
    db.add(wb)
    db.commit()
    db.refresh(wb)
    return _wb_to_dict(wb)


# ─────────────────────────────────────────────────────────────────────────────
# Application FastAPI
# ─────────────────────────────────────────────────────────────────────────────
app = FastAPI(
    title="Nexus Grid — Calculation Engine",
    description=(
        "Moteur de calcul opaque : envoyez vos variables, "
        "recevez vos résultats. Les formules restent protégées."
    ),
    version="1.0.0",
    docs_url="/api/docs",
    openapi_url="/api/openapi.json",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def on_startup():
    init_db()


# Enregistrer le routeur API
app.include_router(api)

# ─────────────────────────────────────────────────────────────────────────────
# Servir le frontend statique (Next.js export) — uniquement si le dossier existe
# ─────────────────────────────────────────────────────────────────────────────
STATIC_DIR = Path(__file__).resolve().parent / "static"

if STATIC_DIR.is_dir():
    # Servir les fichiers statiques Next.js (_next/*, images, etc.)
    _next_dir = STATIC_DIR / "_next"
    if _next_dir.is_dir():
        app.mount("/_next", StaticFiles(directory=_next_dir), name="next-static")

    # Catch-all : pour toute route non-API, servir le fichier HTML correspondant
    # ou index.html comme fallback SPA
    @app.get("/{full_path:path}")
    async def serve_frontend(request: Request, full_path: str):
        # Fichier statique exact (favicon.ico, images, etc.)
        file_path = STATIC_DIR / full_path
        if file_path.is_file():
            return FileResponse(file_path)

        # Page Next.js exportée (ex: /dashboard → dashboard.html)
        html_path = STATIC_DIR / f"{full_path}.html"
        if html_path.is_file():
            return FileResponse(html_path)

        # Sous-dossier avec index.html (ex: /dashboard/ → dashboard/index.html)
        index_path = STATIC_DIR / full_path / "index.html"
        if index_path.is_file():
            return FileResponse(index_path)

        # Fallback → index.html (SPA)
        fallback = STATIC_DIR / "index.html"
        if fallback.is_file():
            return FileResponse(fallback)

        return JSONResponse(status_code=404, content={"detail": "Not found"})
